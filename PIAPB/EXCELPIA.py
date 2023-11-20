import openpyxl

def Excel(titulo, ID, estreno, p_prom):
    nomEx = 'Peliculas_Vistas.xlsx'
    try:
        libro = openpyxl.load_workbook(nomEx)
    except FileNotFoundError:
        libro = openpyxl.Workbook()

    if 'Peliculas Visualizadas' not in libro.sheetnames:
        pag = libro.create_sheet(title='Peliculas Visualizadas')
        encaEX = ["Titulo", "ID", "Fecha Estreno", "Puntuacion"]
        pag.append(encaEX)
    else:
        pag = libro['Peliculas Visualizadas']

    filanu = [titulo, ID, estreno, p_prom]
    pag.append(filanu)

    libro.save(nomEx)
